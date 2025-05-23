import calendar
import os
import datetime
from collections import defaultdict

import pandas as pd
from dateutil.utils import today
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model

from database.DAO import Dao
from database import DB_connect

class SchedulingModel:
    """
    Logica di business per la generazione e gestione dei turni.
    """
    def __init__(self):
        # Configurazione DB, da parametrizzare esternamente

        self.current_year = None
        self.current_month = None
        self.dao = Dao()
        self.turni = []  # lista di turni generati
        self.dictDipendenti = self.dao.get_employees()
        self.is_first_month_historically = False  # Inizializza qui!
        # Mappa i nomi dei turni ai loro tipi principali (M, P, N)




    def controllo_mese(self, mese_str: str) -> bool:
        print(f"DEBUG: controllo_mese chiamato con: {mese_str}")
        try:
            month_val, year_val = map(int, mese_str.split("-"))
            print(f"DEBUG: Parsed month: {month_val}, year: {year_val}")
        except ValueError:
            print("ERRORE: Formato mese non valido. Usa 'MM-YYYY'.")
            self.is_first_month_historically = False # <--- AGGIUNGI ANCHE QUI per sicurezza
            return False

        self.current_month = month_val
        self.current_year = year_val

        # Definisci il primo mese storico (Maggio 2025 secondo il tuo output)
        FIRST_MONTH_YEAR = 2025 # Secondo il tuo output
        FIRST_MONTH_NUM = 5   # Secondo il tuo output

        print(f"DEBUG: Mese corrente impostato a: {self.current_month}-{self.current_year}")
        print(f"DEBUG: Primo mese storico configurato: {FIRST_MONTH_NUM}-{FIRST_MONTH_YEAR}")

        if (self.current_year < FIRST_MONTH_YEAR) or \
           (self.current_year == FIRST_MONTH_YEAR and self.current_month < FIRST_MONTH_NUM):
            print("DEBUG: Condizione TRUE: Mese troppo vecchio")
            print("Mese troppo vecchio per essere il primo del database (anteriormente a Maggio 2025).")
            self.is_first_month_historically = False # <--- IMPOSTA QUI
            return False
        elif self.current_year == FIRST_MONTH_YEAR and self.current_month == FIRST_MONTH_NUM:
            print("DEBUG: Condizione TRUE: Primo mese storico identificato (Maggio 2025)")
            print("Primo mese del database.")
            self.is_first_month_historically = True # <--- IMPOSTA QUI
            self.loadNecessitaDipendenti(mese_str)
            return True
        else:
            print("DEBUG: Condizione: Mese successivo al primo storico.")
            # Controlla l'esistenza del mese precedente usando il DAO
            if self.dao.controlla_mese_precedente(self.current_year, self.current_month):
                print("DEBUG: Mese precedente trovato nel database.")
                self.is_first_month_historically = False # <--- IMPOSTA QUI
                self.loadNecessitaDipendenti(mese_str)
                return True
            else:
                print("DEBUG: Mese precedente NON trovato nel database.")
                print(f"IMPOSSIBILE PROCEDERE: Il mese precedente ({self.current_month-1 if self.current_month > 1 else 12}-{self.current_year if self.current_month > 1 else self.current_year-1}) non è stato trovato nel database. Genera prima i mesi precedenti.")
                self.is_first_month_historically = False # <--- IMPOSTA QUI
                return False




    def loadNecessitaDipendenti(self, mese):
        print("necessita in creazione")
        month_str, year_str = mese.split("-")
        self.current_month = int(month_str)
        self.current_year = int(year_str)
        for dip in self.dictDipendenti.values():
            # Ogni dipendente riceve il suo dict fresco
            nuovo_dict = self.costruttoreDictNecessita()
            dip.setDizionario(nuovo_dict)

    def costruttoreDictNecessita(self) -> dict[str, dict[int, bool]]:
        """
        Costruisce, per un dipendente, il dict con 6 sotto‐dizionari
        giorno→False, usando l'anno e il mese correnti di questa istanza.
        """
        # numero di giorni del mese corrente
        days_in_month = calendar.monthrange(self.current_year, self.current_month)[1]
        days = range(1, days_in_month + 1)

        necessity_types = [
            "permessi_ferie",
            "mutua_infortunio",
            "non_retribuite",
            "no_mattino",
            "no_pomeriggio",
            "no_notte",
        ]

        # dict comprehension
        return {
            ntype: { day: False for day in days }
            for ntype in necessity_types
        }



    def get_employeesM(self):
        return self.dictDipendenti

    def get_employeeM(self, emp_id):
        return self.dictDipendenti[emp_id]


    def updateNrNotti(self, emp_id, nr):
        self.dictDipendenti[emp_id].setEsigenzeNotti(nr)  # CAMBIO NOME METODO
        return True

    def updateNrMattini(self, emp_id, nr):
        self.dictDipendenti[emp_id].setEsigenzeMattini(nr)  # CAMBIO NOME METODO
        return True

    def updateNrPomeriggi(self, emp_id, nr):
        self.dictDipendenti[emp_id].setEsigenzePomeriggi(nr)  # CAMBIO NOME METODO
        return True

    def updateDaIncludere(self, emp_id, tipo):
        # Questi metodi azzeraTurni/ripristinaTurni agiscono sulle esigenze settimanali predefinite
        # e non sui conteggi attuali (che sono gestiti nel Model/generazione turni)
        if tipo == "maternità" or tipo == "aspettativa":
            self.dictDipendenti[emp_id].azzeraTurni()  # Metodo esistente in Dipendente

        else:
            self.dictDipendenti[emp_id].ripristinaTurni()  # Metodo esistente in Dipendente
        self.dictDipendenti[emp_id].setDaIncludere(tipo)





    def set_permesso_dipendente(self, emp_id: int, giorno: int, valore: bool):
        # Assicurati che emp_id sia valido e che il dizionario esista
        if emp_id in self.dictDipendenti:
            if 'permessi_ferie' not in self.dictDipendenti[emp_id].dizionarioNecessita:
                self.dictDipendenti[emp_id].dizionarioNecessita['permessi_ferie'] = {}
            self.dictDipendenti[emp_id].dizionarioNecessita['permessi_ferie'][giorno] = valore
            print(f"DEBUG: set_permesso_dipendente per {emp_id}, giorno {giorno}, valore {valore}. Stato attuale: {self.dictDipendenti[emp_id].dizionarioNecessita['permessi_ferie'][giorno]}")

    def set_mutua_dipendente(self, emp_id: int, giorno: int, valore: bool):
        if emp_id in self.dictDipendenti:
            if 'mutua_infortunio' not in self.dictDipendenti[emp_id].dizionarioNecessita:
                self.dictDipendenti[emp_id].dizionarioNecessita['mutua_infortunio'] = {}
            self.dictDipendenti[emp_id].dizionarioNecessita['mutua_infortunio'][giorno] = valore
            print(f"DEBUG: set_mutua_dipendente per {emp_id}, giorno {giorno}, valore {valore}. Stato attuale: {self.dictDipendenti[emp_id].dizionarioNecessita['mutua_infortunio'][giorno]}")

    def set_esigenza_dipendente(self, emp_id: int, giorno: int, valore: bool):
        # Questo è "non_retribuite"
        if emp_id in self.dictDipendenti:
            if 'non_retribuite' not in self.dictDipendenti[emp_id].dizionarioNecessita:
                self.dictDipendenti[emp_id].dizionarioNecessita['non_retribuite'] = {}
            self.dictDipendenti[emp_id].dizionarioNecessita['non_retribuite'][giorno] = valore
            print(f"DEBUG: set_esigenza_dipendente per {emp_id}, giorno {giorno}, valore {valore}. Stato attuale: {self.dictDipendenti[emp_id].dizionarioNecessita['non_retribuite'][giorno]}")

    def set_pref_noNott_dipendente(self, emp_id: int, giorno: int, valore: bool):
        if emp_id in self.dictDipendenti:
            if 'no_notte' not in self.dictDipendenti[emp_id].dizionarioNecessita:
                self.dictDipendenti[emp_id].dizionarioNecessita['no_notte'] = {}
            self.dictDipendenti[emp_id].dizionarioNecessita['no_notte'][giorno] = valore
            print(f"DEBUG: set_pref_noNott_dipendente per {emp_id}, giorno {giorno}, valore {valore}. Stato attuale: {self.dictDipendenti[emp_id].dizionarioNecessita['no_notte'][giorno]}")

    def set_pref_noMatt_dipendente(self, emp_id: int, giorno: int, valore: bool):
        if emp_id in self.dictDipendenti:
            if 'no_mattino' not in self.dictDipendenti[emp_id].dizionarioNecessita:
                self.dictDipendenti[emp_id].dizionarioNecessita['no_mattino'] = {}
            self.dictDipendenti[emp_id].dizionarioNecessita['no_mattino'][giorno] = valore
            print(f"DEBUG: set_pref_noMatt_dipendente per {emp_id}, giorno {giorno}, valore {valore}. Stato attuale: {self.dictDipendenti[emp_id].dizionarioNecessita['no_mattino'][giorno]}")

    def set_pref_noPom_dipendente(self, emp_id: int, giorno: int, valore: bool):
        if emp_id in self.dictDipendenti:
            if 'no_pomeriggio' not in self.dictDipendenti[emp_id].dizionarioNecessita:
                self.dictDipendenti[emp_id].dizionarioNecessita['no_pomeriggio'] = {}
            self.dictDipendenti[emp_id].dizionarioNecessita['no_pomeriggio'][giorno] = valore
            print(f"DEBUG: set_pref_noPom_dipendente per {emp_id}, giorno {giorno}, valore {valore}. Stato attuale: {self.dictDipendenti[emp_id].dizionarioNecessita['no_pomeriggio'][giorno]}")






    def controlloTuttiCinque(self):
        for dipendente in self.dictDipendenti:
            if dipendente.daIncludere != ("maternità", "aspettativa") and (int(dipendente.notti) + int(dipendente.mattini) + int(dipendente.pomeriggi) < 5):
                #il dipendente con "da Includere" oppure "non selezionato" non ha 5 turni settimanali
                return False
            else:
                return True

    def genera_turni_mese(self):
        """
        Genera i turni mensili usando CP-SAT con vincoli rigidi e preferenze soft.
        """
        import calendar
        import datetime
        from collections import defaultdict
        import pandas as pd
        from openpyxl.styles import PatternFill, Alignment, Font
        from openpyxl.utils import get_column_letter
        from ortools.sat.python import cp_model

        # --- 1) PREPARAZIONE DATI ---
        year, month = self.current_year, self.current_month
        dim = calendar.monthrange(year, month)[1]
        days = [datetime.date(year, month, d) for d in range(1, dim + 1)]

        # Tipi di turno
        shift_types_mattina = ['mat_1', 'mat_2', 'mat_3', 'mat_4']
        shift_types_pomeriggio = ['pom_1', 'pom_2', 'pom_3']
        shift_types_notte = ['notte']
        shift_types_corti = ['mj', 'pc']  # turni da 3 ore
        all_shifts = shift_types_mattina + shift_types_pomeriggio + shift_types_notte + shift_types_corti

        # Dipendenti attivi (esclusi maternità/aspettativa)
        employees = {eid: dip for eid, dip in self.dictDipendenti.items()
                     if dip.daIncludere not in ['MT', 'ASP']}

        if not employees:
            raise RuntimeError("Nessun dipendente attivo per generare turni")

        # Raggruppa giorni per settimana ISO
        weeks = defaultdict(list)
        for d in days:
            week_num = d.isocalendar()[1]
            weeks[week_num].append(d)

        print(f"Generazione turni per {len(employees)} dipendenti, {dim} giorni, {len(weeks)} settimane")

        # --- 2) CREAZIONE MODELLO CP-SAT ---
        model = cp_model.CpModel()

        # Variabili principali
        x = {}  # x[(emp_id, day, shift)] = 1 se dipendente fa quel turno quel giorno
        has_rest = {}  # has_rest[(emp_id, day)] = 1 se dipendente riposa quel giorno
        has_vacation = {}  # has_vacation[(emp_id, day)] = 1 se dipendente ha ferie/mutua/nr

        # Creazione variabili
        for emp_id in employees:
            for day in days:
                # Variabile riposo
                has_rest[(emp_id, day)] = model.NewBoolVar(f'rest_{emp_id}_{day.day}')

                # Variabile ferie/mutua/non_retribuite
                has_vacation[(emp_id, day)] = model.NewBoolVar(f'vacation_{emp_id}_{day.day}')

                # Variabili turni lavorativi
                for shift in all_shifts:
                    x[(emp_id, day, shift)] = model.NewBoolVar(f'work_{emp_id}_{day.day}_{shift}')

        # --- 3) VINCOLI RIGIDI ---

        # 3.1) Esclusività giornaliera: ogni dipendente fa esattamente una cosa al giorno
        for emp_id in employees:
            for day in days:
                all_day_vars = [has_rest[(emp_id, day)], has_vacation[(emp_id, day)]]
                all_day_vars.extend([x[(emp_id, day, shift)] for shift in all_shifts])
                model.AddExactlyOne(all_day_vars)

        # 3.2) Copertura turni: ogni turno ogni giorno deve essere coperto da esattamente 1 persona
        for day in days:
            for shift in all_shifts:
                shift_vars = [x[(emp_id, day, shift)] for emp_id in employees]
                model.AddExactlyOne(shift_vars)
#---------------------------------------------------------------------------------------------------------------------------------------------
        # 3.3) Esattamente 2 riposi a settimana per ogni dipendente
        for emp_id in employees:
            for week_days in weeks.values():
                if week_days:  # Controlla che la settimana non sia vuota
                    rest_vars = [has_rest[(emp_id, day)] for day in week_days]
                    model.Add(sum(rest_vars) >= 2)
                    model.Add(sum(rest_vars) <= 2)
#---------------------------------------------------------------------------------------------------------------------------------------------------
        # 3.4) Necessità obbligatorie dal dizionarioNecessita
        for emp_id, dip in employees.items():
            if dip.dizionarioNecessita:
                for day in days:
                    day_num = day.day

                    # Ferie obbligatorie
                    if dip.dizionarioNecessita.get('permessi_ferie', {}).get(day_num, False):
                        model.Add(has_vacation[(emp_id, day)] == 1)

                    # Mutua obbligatoria
                    if dip.dizionarioNecessita.get('mutua_infortunio', {}).get(day_num, False):
                        model.Add(has_vacation[(emp_id, day)] == 1)

                    # Non retribuite obbligatorie
                    if dip.dizionarioNecessita.get('non_retribuite', {}).get(day_num, False):
                        model.Add(has_vacation[(emp_id, day)] == 1)

        # --- 4) VINCOLI SOFT (PREFERENZE) ---

        # 4.1) Preferenze per quote settimanali
        preference_violations = []

        for emp_id, dip in employees.items():
            target_mattini = int(dip.esigenzeMattiniSettimanali)
            target_pomeriggi = int(dip.esigenzePomeriggiSettimanali)
            target_notti = int(dip.esigenzeNottiSettimanali)

            for week_days in weeks.values():
                if not week_days:
                    continue

                # Conta turni effettivi per tipo nella settimana
                mattini_vars = [x[(emp_id, day, shift)]
                                for day in week_days for shift in shift_types_mattina]
                pomeriggi_vars = [x[(emp_id, day, shift)]
                                  for day in week_days for shift in shift_types_pomeriggio]
                notti_vars = [x[(emp_id, day, shift)]
                              for day in week_days for shift in shift_types_notte]

                if mattini_vars and target_mattini > 0:
                    # Variabile di violazione per mattini
                    under_mattini = model.NewIntVar(0, 10, f'under_mattini_{emp_id}_{week_days[0].isocalendar()[1]}')
                    over_mattini = model.NewIntVar(0, 10, f'over_mattini_{emp_id}_{week_days[0].isocalendar()[1]}')
                    model.Add(sum(mattini_vars) + under_mattini - over_mattini == target_mattini)
                    preference_violations.extend([under_mattini, over_mattini])

                if pomeriggi_vars and target_pomeriggi > 0:
                    # Variabile di violazione per pomeriggi
                    under_pom = model.NewIntVar(0, 10, f'under_pom_{emp_id}_{week_days[0].isocalendar()[1]}')
                    over_pom = model.NewIntVar(0, 10, f'over_pom_{emp_id}_{week_days[0].isocalendar()[1]}')
                    model.Add(sum(pomeriggi_vars) + under_pom - over_pom == target_pomeriggi)
                    preference_violations.extend([under_pom, over_pom])

                if notti_vars and target_notti > 0:
                    # Variabile di violazione per notti
                    under_notti = model.NewIntVar(0, 10, f'under_notti_{emp_id}_{week_days[0].isocalendar()[1]}')
                    over_notti = model.NewIntVar(0, 10, f'over_notti_{emp_id}_{week_days[0].isocalendar()[1]}')
                    model.Add(sum(notti_vars) + under_notti - over_notti == target_notti)
                    preference_violations.extend([under_notti, over_notti])

        # 4.2) Preferenze negative (no_mattino, no_pomeriggio, no_notte)
        negative_violations = []

        for emp_id, dip in employees.items():
            if dip.dizionarioNecessita:
                for day in days:
                    day_num = day.day

                    # No mattino
                    if dip.dizionarioNecessita.get('no_mattino', {}).get(day_num, False):
                        for shift in shift_types_mattina:
                            violation = model.NewBoolVar(f'no_matt_viol_{emp_id}_{day_num}_{shift}')
                            model.Add(x[(emp_id, day, shift)] <= violation)
                            negative_violations.append(violation)

                    # No pomeriggio
                    if dip.dizionarioNecessita.get('no_pomeriggio', {}).get(day_num, False):
                        for shift in shift_types_pomeriggio:
                            violation = model.NewBoolVar(f'no_pom_viol_{emp_id}_{day_num}_{shift}')
                            model.Add(x[(emp_id, day, shift)] <= violation)
                            negative_violations.append(violation)

                    # No notte
                    if dip.dizionarioNecessita.get('no_notte', {}).get(day_num, False):
                        for shift in shift_types_notte:
                            violation = model.NewBoolVar(f'no_notte_viol_{emp_id}_{day_num}_{shift}')
                            model.Add(x[(emp_id, day, shift)] <= violation)
                            negative_violations.append(violation)

        # Obiettivo: minimizzare violazioni delle preferenze
        total_violations = preference_violations + negative_violations
        if total_violations:
            model.Minimize(sum(total_violations))

        # --- 5) RISOLUZIONE ---
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 60.0  # Timeout di 60 secondi

        print("Risoluzione del modello in corso...")
        status = solver.Solve(model)

        if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            raise RuntimeError(f"Nessuna soluzione trovata. Status: {solver.StatusName(status)}")

        print(f"Soluzione trovata! Status: {solver.StatusName(status)}")
        if total_violations:
            print(f"Violazioni preferenze: {solver.ObjectiveValue()}")

        # --- 6) ESTRAZIONE RISULTATI ---
        db_schedule = []
        excel_rows = []

        for emp_id, dip in employees.items():
            name = f"{dip.nome} {dip.cognome} (ID:{emp_id})"

            for day in days:
                day_num = day.day
                assigned_shift = None
                shift_type_for_db = None
                ore_assegnate = 0
                note = ""

                # Determina cosa fa il dipendente questo giorno
                if solver.Value(has_rest[(emp_id, day)]) == 1:
                    assigned_shift = "RIPOSO"
                    shift_type_for_db = "RIPOSO"
                    ore_assegnate = 0

                elif solver.Value(has_vacation[(emp_id, day)]) == 1:
                    # Determina il tipo specifico di assenza
                    if dip.dizionarioNecessita:
                        if dip.dizionarioNecessita.get('permessi_ferie', {}).get(day_num, False):
                            assigned_shift = "FERIE"
                            shift_type_for_db = "FERIE"
                            ore_assegnate = 6.4  # 80% di 8 ore
                        elif dip.dizionarioNecessita.get('mutua_infortunio', {}).get(day_num, False):
                            assigned_shift = "MUTUA"
                            shift_type_for_db = "MUTUA"
                            ore_assegnate = 6.4  # 80% di 8 ore
                        elif dip.dizionarioNecessita.get('non_retribuite', {}).get(day_num, False):
                            assigned_shift = "NR"
                            shift_type_for_db = "NR"
                            ore_assegnate = 0
                        else:
                            # Ferie automatiche (non dovrebbe succedere con i nuovi vincoli)
                            assigned_shift = "FERIE"
                            shift_type_for_db = "FERIE"
                            ore_assegnate = 6.4
                            note = "Ferie auto"
                    else:
                        assigned_shift = "FERIE"
                        shift_type_for_db = "FERIE"
                        ore_assegnate = 6.4
                        note = "Ferie auto"
                else:
                    # Turno lavorativo
                    for shift in all_shifts:
                        if solver.Value(x[(emp_id, day, shift)]) == 1:
                            assigned_shift = shift
                            shift_type_for_db = shift
                            ore_assegnate = 8 if shift not in shift_types_corti else 3
                            break

                    if assigned_shift is None:
                        raise RuntimeError(f"Errore: dipendente {emp_id} giorno {day_num} senza assegnazione")

                # Aggiungi ai risultati
                db_schedule.append({
                    'data_turno': day.isoformat(),
                    'codice_dipendente': emp_id,
                    'tipo_turno': shift_type_for_db,
                    'ore_assegnate': ore_assegnate,
                    'note': note
                })

                excel_rows.append({
                    'Dipendente': name,
                    'Giorno': day_num,
                    'Turno': assigned_shift
                })

        # --- 7) SALVATAGGIO DATABASE ---
        print("Salvataggio turni nel database...")
        self.dao.save_turni_mese(year, month, db_schedule)

        # --- 8) GENERAZIONE EXCEL ---
        print("Generazione file Excel...")
        df = pd.DataFrame(excel_rows)
        pivot = df.pivot(index='Dipendente', columns='Giorno', values='Turno')

        filename = f"turni_{year}_{month:02d}.xlsx"
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            pivot.to_excel(writer, sheet_name='Turni')
            wb = writer.book
            ws = writer.sheets['Turni']

            # Formattazione intestazioni
            for cell in ws[1]:
                if cell.value:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

            # Mappa colori per tipo turno
            color_map = {
                'mat_': 'FFF2CC',  # Giallo chiaro per mattina
                'pom_': 'D9EAD3',  # Verde chiaro per pomeriggio
                'notte': 'CFE2F3',  # Azzurro chiaro per notte
                'mj': 'F4CCCC',  # Rosa chiaro per mj
                'pc': 'EAD1DC',  # Viola chiaro per pc
                'RIPOSO': 'FFFFFF',  # Bianco per riposo
                'FERIE': 'FCE5CD',  # Arancione chiaro per ferie
                'MUTUA': 'D0E0E3',  # Grigio azzurro per mutua
                'NR': 'F4CCCC'  # Rosa per non retribuite
            }

            # Applica colori
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=1 + dim):
                for cell in row:
                    if cell.value:
                        val = str(cell.value)
                        for prefix, color in color_map.items():
                            if val.startswith(prefix) or val == prefix:
                                cell.fill = PatternFill("solid", fgColor=color)
                                cell.alignment = Alignment(horizontal='center')
                                break

            # Regola larghezza colonne
            for idx, column_cells in enumerate(ws.columns, 1):
                if column_cells:
                    max_length = max(len(str(cell.value or '')) for cell in column_cells)
                    adjusted_width = min(max_length + 2, 15)  # Max 15 caratteri
                    ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

        print(f"✅ Turni generati con successo!")
        print(f"📊 File Excel: {filename}")
        print(f"💾 Database aggiornato per {year}-{month:02d}")

        return filename